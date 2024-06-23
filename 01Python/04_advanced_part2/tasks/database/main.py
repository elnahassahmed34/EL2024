import openpyxl
from openpyxl import Workbook

# Function to create the Excel sheet if it doesn't exist
def create_excel_file(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employees"
        sheet.append(["ID", "Name", "Job", "Salary"])
        workbook.save(file_name)
    return workbook

# Function to generate a new unique ID
def generate_id(sheet):
    max_row = sheet.max_row
    if max_row == 1:
        return "IDX001"
    else:
        last_id = sheet.cell(row=max_row, column=1).value
        new_id_num = int(last_id[3:]) + 1
        new_id = f"IDX{new_id_num:03}"
        return new_id

# Function to add a new employee
def add_employee(sheet, name, job, salary):
    new_id = generate_id(sheet)
    sheet.append([new_id, name, job, salary])
    print(f"Employee '{name}' added successfully with ID {new_id}.")

# Function to print employee data
def print_employee_data(sheet):
    if sheet.max_row == 1:
        print("No employees found.")
    else:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(f"ID: {row[0]}, Name: {row[1]}, Job: {row[2]}, Salary: {row[3]}")

# Function to remove an employee
def remove_employee(sheet, employee_id):
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == employee_id:
            sheet.delete_rows(row[0].row)
            print(f"Employee with ID '{employee_id}' removed successfully.")
            return
    print(f"No employee found with ID '{employee_id}'.")

# Function to update an employee's data
def update_employee(sheet, employee_id, name=None, job=None, salary=None):
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == employee_id:
            if name:
                row[1].value = name
            if job:
                row[2].value = job
            if salary:
                row[3].value = salary
            print(f"Employee with ID '{employee_id}' updated successfully.")
            return
    print(f"No employee found with ID '{employee_id}'.")

# Function to display the menu
def display_menu():
    print("\nEmployee Management System")
    print("1. Add new employee")
    print("2. Print employee data")
    print("3. Remove employee")
    print("4. Update employee")
    print("5. Exit")

if __name__ == "__main__":
    file_name = "employees.xlsx"
    workbook = create_excel_file(file_name)
    sheet = workbook["Employees"]

    while True:
        display_menu()
        choice = input("Enter your choice: ")

        if choice == '1':
            name = input("Enter employee name: ")
            job = input("Enter employee job: ")
            salary = float(input("Enter employee salary: "))
            add_employee(sheet, name, job, salary)

        elif choice == '2':
            print_employee_data(sheet)

        elif choice == '3':
            employee_id = input("Enter employee ID to remove: ")
            remove_employee(sheet, employee_id)

        elif choice == '4':
            employee_id = input("Enter employee ID to update: ")
            name = input("Enter new name (leave blank to skip): ")
            job = input("Enter new job (leave blank to skip): ")
            salary = input("Enter new salary (leave blank to skip): ")
            update_employee(sheet, employee_id, name=name or None, job=job or None, salary=float(salary) if salary else None)

        elif choice == '5':
            print("Exiting the system.")
            break

        else:
            print("Invalid choice. Please try again.")

        workbook.save(file_name)
